from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


QUESTION_TYPES = {
    "single": "单选题",
    "multi": "多选题",
    "judge": "判断题",
    "fill": "填空题",
    "short": "简答题",
}


COLUMNS = [
    {"key": "type", "name": "题型", "width": 12, "required": True, "hint": "single/multi/judge/fill/short"},
    {"key": "content", "name": "题干", "width": 60, "required": True, "hint": "题目内容"},
    {"key": "option_a", "name": "选项A", "width": 30, "required": False, "hint": "选择题填写"},
    {"key": "option_b", "name": "选项B", "width": 30, "required": False, "hint": "选择题填写"},
    {"key": "option_c", "name": "选项C", "width": 30, "required": False, "hint": "选择题填写"},
    {"key": "option_d", "name": "选项D", "width": 30, "required": False, "hint": "选择题填写"},
    {"key": "option_e", "name": "选项E", "width": 30, "required": False, "hint": "多选题可填"},
    {"key": "option_f", "name": "选项F", "width": 30, "required": False, "hint": "多选题可填"},
    {"key": "answer", "name": "答案", "width": 20, "required": True, "hint": "单选填A/B/C/D，多选填ABCD，判断填T/F"},
    {"key": "explanation", "name": "解析", "width": 50, "required": False, "hint": "题目解析（选填）"},
]


def generate_excel_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "题目导入模板"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    hint_font = Font(name="微软雅黑", size=9, color="999999")
    hint_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=True)
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col["name"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]

    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=col["hint"])
        cell.font = hint_font
        cell.fill = hint_fill
        cell.alignment = data_alignment
        cell.border = data_border

    examples = [
        {
            "type": "single",
            "content": "下列哪个选项是正确的？",
            "option_a": "选项A内容",
            "option_b": "选项B内容",
            "option_c": "选项C内容（正确答案）",
            "option_d": "选项D内容",
            "answer": "C",
            "explanation": "本题考查基础概念，根据知识点X，C选项是正确答案。",
        },
        {
            "type": "multi",
            "content": "下列哪些选项是正确的？（多选）",
            "option_a": "选项A（正确）",
            "option_b": "选项B（正确）",
            "option_c": "选项C（错误）",
            "option_d": "选项D（正确）",
            "answer": "ABD",
            "explanation": "本题考查多个知识点，ABD三个选项都符合条件。",
        },
        {
            "type": "judge",
            "content": "Python是一种解释型编程语言。",
            "answer": "T",
            "explanation": "Python确实是解释型语言，不需要编译即可运行。",
        },
        {
            "type": "fill",
            "content": "TCP/IP协议中，TCP是______层协议，IP是______层协议。",
            "answer": "传输；网络",
            "explanation": "TCP属于传输层，负责可靠传输；IP属于网络层，负责路由和寻址。",
        },
        {
            "type": "short",
            "content": "请简述什么是面向对象编程及其三大特性。",
            "answer": "面向对象编程是一种编程范式，将数据和操作数据的方法封装在一起。三大特性：封装、继承、多态。",
            "explanation": "封装隐藏内部实现，继承实现代码复用，多态实现接口统一。",
        },
    ]

    for row_idx, example in enumerate(examples, 3):
        for col_idx, col in enumerate(COLUMNS, 1):
            key = col["key"]
            value = example.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border

    ws.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_excel_to_questions(file_bytes: bytes) -> list:
    from openpyxl import load_workbook
    from io import BytesIO

    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    questions = []
    errors = []

    for row_idx in range(3, ws.max_row + 1):
        try:
            row_data = {}
            for col_idx, col in enumerate(COLUMNS, 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data[col["key"]] = str(cell_value).strip() if cell_value else ""

            q_type = row_data["type"]
            content = row_data["content"]

            if not content:
                continue

            if q_type not in QUESTION_TYPES:
                errors.append(f"第{row_idx}行：题型 '{q_type}' 无效，应为 single/multi/judge/fill/short")
                continue

            options = {}
            if q_type in ["single", "multi"]:
                for opt_key in ["option_a", "option_b", "option_c", "option_d", "option_e", "option_f"]:
                    opt_value = row_data[opt_key]
                    if opt_value:
                        options[opt_key.replace("option_", "").upper()] = opt_value

            elif q_type == "judge":
                options = {"T": "正确", "F": "错误"}

            answer = row_data["answer"]
            explanation = row_data["explanation"]

            if not answer and q_type not in ["short", "fill"]:
                errors.append(f"第{row_idx}行：答案不能为空")
                continue

            questions.append({
                "type": q_type,
                "content": content,
                "options": options,
                "answer": answer,
                "explanation": explanation,
            })
        except Exception as e:
            errors.append(f"第{row_idx}行：解析错误 - {str(e)}")

    return questions, errors


def parse_csv_to_questions(file_bytes: bytes, encoding: str = "utf-8-sig") -> list:
    import csv
    from io import StringIO

    try:
        text = file_bytes.decode(encoding)
    except UnicodeDecodeError:
        text = file_bytes.decode("gbk")

    reader = csv.DictReader(StringIO(text))
    questions = []
    errors = []
    row_idx = 2

    for row in reader:
        try:
            q_type = row.get("题型", "").strip() or row.get("type", "").strip()
            content = row.get("题干", "").strip() or row.get("content", "").strip()

            if not content:
                row_idx += 1
                continue

            if q_type not in QUESTION_TYPES and q_type:
                errors.append(f"第{row_idx}行：题型 '{q_type}' 无效")
                row_idx += 1
                continue

            if not q_type:
                q_type = "single"

            options = {}
            if q_type in ["single", "multi"]:
                for key in ["A", "B", "C", "D", "E", "F"]:
                    opt_value = row.get(f"选项{key}", "").strip() or row.get(f"option_{key.lower()}", "").strip() or row.get(f"option{key.lower()}", "").strip()
                    if opt_value:
                        options[key] = opt_value

            elif q_type == "judge":
                options = {"T": "正确", "F": "错误"}

            answer = row.get("答案", "").strip() or row.get("answer", "").strip()
            explanation = row.get("解析", "").strip() or row.get("explanation", "").strip()

            if not answer and q_type not in ["short", "fill"]:
                errors.append(f"第{row_idx}行：答案不能为空")
                row_idx += 1
                continue

            questions.append({
                "type": q_type,
                "content": content,
                "options": options,
                "answer": answer,
                "explanation": explanation,
            })
        except Exception as e:
            errors.append(f"第{row_idx}行：解析错误 - {str(e)}")

        row_idx += 1

    return questions, errors


def parse_json_to_questions(file_bytes: bytes) -> list:
    import json

    try:
        data = json.loads(file_bytes)
    except json.JSONDecodeError as e:
        return [], [f"JSON解析错误: {str(e)}"]

    questions = []
    errors = []

    if isinstance(data, dict):
        data = data.get("questions", data.get("data", []))

    if not isinstance(data, list):
        return [], ["JSON格式错误：根节点应为数组或包含questions字段的对象"]

    for idx, item in enumerate(data, 1):
        try:
            q_type = item.get("type", "single")
            content = item.get("content", "").strip()

            if not content:
                errors.append(f"第{idx}条：题干不能为空")
                continue

            if q_type not in QUESTION_TYPES:
                errors.append(f"第{idx}条：题型 '{q_type}' 无效")
                continue

            options = item.get("options", {})
            if isinstance(options, dict):
                normalized_options = {}
                for key, value in options.items():
                    normalized_options[str(key).upper()] = str(value).strip()
                options = normalized_options

            elif q_type == "judge" and not options:
                options = {"T": "正确", "F": "错误"}

            answer = str(item.get("answer", "")).strip()
            explanation = str(item.get("explanation", "")).strip()

            if not answer and q_type not in ["short", "fill"]:
                errors.append(f"第{idx}条：答案不能为空")
                continue

            questions.append({
                "type": q_type,
                "content": content,
                "options": options,
                "answer": answer,
                "explanation": explanation,
            })
        except Exception as e:
            errors.append(f"第{idx}条：解析错误 - {str(e)}")

    return questions, errors