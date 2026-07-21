import base64
import json
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import httpx

from utils.answer_normalizer import normalize_question_type, normalize_standard_answer
from utils.file_extract import IMAGE_EXTENSIONS, TEXT_EXTRACT_EXTENSIONS, _extract_from_txt, extract_text_from_file

DEEPSEEK_API_URL = "https://api.deepseek.com/v1/chat/completions"
DEEPSEEK_MODEL = "deepseek-v4-flash"
VALID_OPTION_KEYS = ("A", "B", "C", "D", "E", "F")


def _strip_code_fence(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("```"):
        stripped = re.sub(r"^```[a-zA-Z0-9_-]*\n?", "", stripped)
        stripped = re.sub(r"\n?```$", "", stripped)
    return stripped.strip()


def _extract_json_payload(text: str) -> Any:
    cleaned = _strip_code_fence(text)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    match = re.search(r"(\[[\s\S]*\]|\{[\s\S]*\})", cleaned)
    if not match:
        raise ValueError("AI 未返回可解析的 JSON")
    return json.loads(match.group(1))


def _normalize_options(raw_options: Any, question_type: str) -> dict[str, str]:
    if question_type == "judge":
        return {"T": "正确", "F": "错误"}

    if not isinstance(raw_options, dict):
        return {}

    normalized: dict[str, str] = {}
    for key in VALID_OPTION_KEYS:
        value = raw_options.get(key)
        if value is None:
            value = raw_options.get(key.lower())
        if value is None:
            continue
        text = str(value).strip()
        if text:
            normalized[key] = text
    return normalized


def _validate_question_payload(question: dict[str, Any], index: int) -> tuple[Optional[dict[str, Any]], Optional[str]]:
    question_type = normalize_question_type(question.get("type", "single"))
    content = str(question.get("content", "")).strip()
    explanation = str(question.get("explanation", "")).strip()
    options = _normalize_options(question.get("options", {}), question_type)
    answer = normalize_standard_answer(question_type, str(question.get("answer", "")).strip())

    if not content:
        return None, f"第 {index} 题缺少题干"

    if question_type == "single":
        if len(options) < 2:
            return None, f"第 {index} 题单选题选项不足"
        if not re.fullmatch(r"[A-F]", answer) or answer not in options:
            return None, f"第 {index} 题单选题答案无效"
    elif question_type == "multi":
        if len(options) < 2:
            return None, f"第 {index} 题多选题选项不足"
        letters = list(answer)
        if not letters or any(letter not in options for letter in letters):
            return None, f"第 {index} 题多选题答案无效"
    elif question_type == "judge":
        if answer not in {"T", "F"}:
            return None, f"第 {index} 题判断题答案无效"
    elif question_type == "fill":
        if not answer:
            return None, f"第 {index} 题填空题缺少答案"
    elif question_type in {"short", "code"}:
        pass

    return {
        "type": question_type,
        "content": content,
        "options": options,
        "answer": answer,
        "explanation": explanation,
    }, None


def sanitize_ai_questions(payload: Any) -> tuple[list[dict[str, Any]], list[str]]:
    if isinstance(payload, dict):
        items = payload.get("questions", [])
    else:
        items = payload

    if not isinstance(items, list):
        return [], ["AI 返回的数据不是题目数组"]

    questions: list[dict[str, Any]] = []
    errors: list[str] = []
    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            errors.append(f"第 {index} 题数据格式无效")
            continue
        normalized, error = _validate_question_payload(item, index)
        if error:
            errors.append(error)
            continue
        questions.append(normalized)

    return questions, errors


def extract_text_for_ai_fallback(filename: str, file_bytes: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in TEXT_EXTRACT_EXTENSIONS:
        return extract_text_from_file(filename, file_bytes)
    if suffix == ".csv":
        return _extract_from_txt(file_bytes)
    if suffix == ".json":
        return file_bytes.decode("utf-8")
    if suffix in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            raise ValueError(f"服务器缺少 Excel 解析依赖: {exc}") from exc

        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        worksheet = workbook.active
        rows: list[str] = []
        for row in worksheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                rows.append("\t".join(values))
        return "\n".join(rows).strip()
    raise ValueError(f"暂不支持对 {suffix or '无扩展名文件'} 进行 AI 兜底解析")


async def parse_questions_with_ai(
    text: str,
    source_name: Optional[str] = None,
    api_key: Optional[str] = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    if not api_key:
        return [], ["尚未配置个人 DeepSeek API Key"]

    cleaned_text = str(text or "").strip()
    if not cleaned_text:
        return [], ["缺少可供 AI 解析的文本内容"]

    prompt = f"""你是题库导入助手。请把下面的原始内容提取成题目 JSON。
要求：1. 只能输出 JSON，不能输出任何额外解释。2. 输出格式必须是数组，每个元素包含：type, content, options, answer, explanation。3. type 只能是：single, multi, judge, fill, short, code。4. single/multi 的 options 必须是对象，键只能是 A-F。5. judge 的 answer 只能是 T 或 F，options 固定为 {{"T":"正确","F":"错误"}}。6. fill/short/code 没有选项时，options 返回空对象。7. 无法确定的题目不要编造，直接跳过。8. 如果原文里没有解析，explanation 返回空字符串。
来源：{source_name or '用户导入内容'}
原始内容如下：
{cleaned_text}
"""

    try:
        async with httpx.AsyncClient(timeout=90.0) as client:
            response = await client.post(
                DEEPSEEK_API_URL,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": DEEPSEEK_MODEL,
                    "messages": [
                        {"role": "system", "content": "你是一个只返回合法 JSON 的题库结构化提取器。"},
                        {"role": "user", "content": prompt},
                    ],
                    "temperature": 0.1,
                    "max_tokens": 4000,
                    "response_format": {"type": "json_object"},
                },
            )
    except httpx.TimeoutException:
        return [], ["AI 解析超时，请稍后重试"]
    except httpx.ConnectError:
        return [], ["无法连接 DeepSeek API，请检查网络"]
    except Exception as exc:
        return [], [f"AI 解析请求失败: {exc}"]

    if response.status_code == 401:
        return [], ["DeepSeek API Key 无效或已过期"]
    if response.status_code == 429:
        return [], ["DeepSeek API 调用次数过多，请稍后重试"]
    if response.status_code >= 400:
        return [], [f"DeepSeek API 调用失败: HTTP {response.status_code}"]

    try:
        result = response.json()
        content = result["choices"][0]["message"]["content"]
        payload = _extract_json_payload(content)
    except Exception as exc:
        return [], [f"AI 返回结果无法解析: {exc}"]

    return sanitize_ai_questions(payload)


async def parse_image_with_ai(
    image_bytes: bytes,
    filename: str,
    api_key: Optional[str] = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    if not api_key:
        return [], ["尚未配置个人 DeepSeek API Key"]

    if not image_bytes:
        return [], ["图片文件为空"]

    try:
        image_base64 = base64.b64encode(image_bytes).decode("utf-8")
    except Exception as exc:
        return [], [f"图片编码失败: {exc}"]

    suffix = Path(filename).suffix.lower().lstrip(".")
    if not suffix:
        suffix = "png"

    prompt = """你是题库导入助手。请识别图片中的题目内容，并提取成题目 JSON。
要求：1. 只能输出 JSON，不能输出任何额外解释。2. 输出格式必须是数组，每个元素包含：type, content, options, answer, explanation。3. type 只能是：single, multi, judge, fill, short, code。4. single/multi 的 options 必须是对象，键只能是 A-F。5. judge 的 answer 只能是 T 或 F，options 固定为 {"T":"正确","F":"错误"}。6. fill/short/code 没有选项时，options 返回空对象。7. 无法确定的题目不要编造，直接跳过。8. 如果图片里没有解析，explanation 返回空字符串。9. 如果图片中有多道题目，请全部识别并输出。"""

    try:
        async with httpx.AsyncClient(timeout=90.0) as client:
            response = await client.post(
                DEEPSEEK_API_URL,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": "deepseek-v2-chat",
                    "messages": [
                        {"role": "system", "content": "你是一个只返回合法 JSON 的题库结构化提取器。"},
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:image/{suffix};base64,{image_base64}",
                                    },
                                },
                            ],
                        },
                    ],
                    "temperature": 0.1,
                    "max_tokens": 4000,
                    "response_format": {"type": "json_object"},
                },
            )
    except httpx.TimeoutException:
        return [], ["AI 解析超时，请稍后重试"]
    except httpx.ConnectError:
        return [], ["无法连接 DeepSeek API，请检查网络"]
    except Exception as exc:
        return [], [f"AI 解析请求失败: {exc}"]

    if response.status_code == 401:
        return [], ["DeepSeek API Key 无效或已过期"]
    if response.status_code == 429:
        return [], ["DeepSeek API 调用次数过多，请稍后重试"]
    if response.status_code >= 400:
        return [], [f"DeepSeek API 调用失败: HTTP {response.status_code}"]

    try:
        result = response.json()
        content = result["choices"][0]["message"]["content"]
        payload = _extract_json_payload(content)
    except Exception as exc:
        return [], [f"AI 返回结果无法解析: {exc}"]

    return sanitize_ai_questions(payload)
